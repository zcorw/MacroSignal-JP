from __future__ import annotations

import calendar
import logging
import math
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from src.clean_data import Observation

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ForeignResidentObservation:
    date: str
    year: int
    month: int
    nationality: str
    nationality_code: str | None
    residence_status: str
    residence_status_code: str | None
    prefecture: str
    prefecture_code: str | None
    gender: str | None
    age_group: str | None
    age: str | None
    value: int
    unit: str
    source_name: str
    source_file: str


@dataclass(frozen=True)
class ForeignResidentMetric:
    date: str
    metric_key: str
    dimension: str
    dimension_value: str
    value: float | None
    unit: str
    source_name: str
    source_file: str


@dataclass(frozen=True)
class ForeignWorkerMetric:
    date: str
    metric_key: str
    dimension: str
    dimension_value: str
    value: float | None
    unit: str
    source_name: str
    source_file: str


@dataclass(frozen=True)
class ForeignWageMetric:
    date: str
    metric_key: str
    dimension: str
    dimension_value: str
    value: float | None
    unit: str
    source_name: str
    source_file: str


LONG_TERM_STATUS_KEYWORDS = (
    "永住者",
    "定住者",
    "日本人の配偶者等",
    "永住者の配偶者等",
)


def clean_foreign_residents(
    raw_dir: Path,
    macro_observations: list[Observation] | None = None,
) -> tuple[
    list[ForeignResidentObservation],
    list[ForeignResidentMetric],
    list[ForeignWorkerMetric],
    list[ForeignWageMetric],
]:
    source_dir = raw_dir / "foreign_residents"
    if not source_dir.exists():
        return [], [], [], []

    files = sorted(source_dir.glob("estat_foreign_residents*.xlsx"))
    resident_metrics: list[ForeignResidentMetric] = []
    for path in files:
        logger.info("开始聚合清洗在留外国人文件：%s", path)
        resident_metrics.extend(clean_foreign_resident_metrics_file(path))
        logger.info("在留外国人指标累计 %d 条", len(resident_metrics))
    resident_metrics.extend(calculate_foreign_resident_change_metrics(resident_metrics))

    logger.info("开始清洗外国人劳动者指标")
    worker_metrics = clean_foreign_worker_metrics(source_dir)
    logger.info("外国人劳动者指标清洗完成：%d 条", len(worker_metrics))

    logger.info("开始清洗外国人工资指标")
    wage_metrics = clean_foreign_wage_metrics(source_dir, macro_observations or [])
    logger.info("外国人工资指标清洗完成：%d 条", len(wage_metrics))
    return [], resident_metrics, worker_metrics, wage_metrics


def clean_foreign_resident_metrics_file(path: Path) -> list[ForeignResidentMetric]:
    required = {"国籍・地域", "在留資格", "都道府県", "在留外国人数"}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        detail_sheet = find_detail_sheet(workbook.sheetnames)
        if detail_sheet is None:
            return []

        year, month = parse_period_from_sheet(detail_sheet)
        period_date = month_end(year, month)
        sheet = workbook[detail_sheet]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []

        column_index = {clean_text(name): index for index, name in enumerate(header) if clean_text(name)}
        if not required.issubset(set(column_index)):
            logger.warning("在留外国人文件缺少必要列：%s", path)
            return []

        nationality_totals: dict[str, float] = {}
        status_totals: dict[str, float] = {}
        total = 0.0
        long_term_total = 0.0
        row_count = 0
        valid_count = 0

        for row in rows:
            row_count += 1
            value = parse_int(row[column_index["在留外国人数"]])
            if value is None:
                continue

            _, nationality = split_code_label(row[column_index["国籍・地域"]])
            _, residence_status = split_code_label(row[column_index["在留資格"]])
            if not nationality or not residence_status:
                continue

            numeric_value = float(value)
            total += numeric_value
            valid_count += 1
            nationality_totals[nationality] = nationality_totals.get(nationality, 0.0) + numeric_value
            status_totals[residence_status] = status_totals.get(residence_status, 0.0) + numeric_value
            if is_long_term_status(residence_status):
                long_term_total += numeric_value

            if row_count % 100000 == 0:
                logger.info("在留外国人流式聚合进度 sheet=%s：已扫描 %d 行", detail_sheet, row_count)

        logger.info("在留外国人流式聚合完成 sheet=%s：扫描 %d 行，有效 %d 行", detail_sheet, row_count, valid_count)
        return calculate_foreign_resident_metrics_from_aggregates(
            date_value=period_date,
            source_file=str(path),
            total=total,
            nationality_totals=nationality_totals,
            status_totals=status_totals,
            long_term_total=long_term_total,
        )
    finally:
        workbook.close()


def calculate_foreign_resident_metrics_from_aggregates(
    date_value: str,
    source_file: str,
    total: float,
    nationality_totals: dict[str, float],
    status_totals: dict[str, float],
    long_term_total: float,
) -> list[ForeignResidentMetric]:
    if not total:
        return []

    source_name = "e-Stat Foreign Residents"
    metrics = [
        ForeignResidentMetric(
            date=date_value,
            metric_key="foreign_residents_total",
            dimension="total",
            dimension_value="all",
            value=total,
            unit="persons",
            source_name=source_name,
            source_file=source_file,
        )
    ]
    metrics.extend(
        calculate_top_aggregate_share_metrics(
            date_value,
            "nationality",
            "foreign_residents_by_nationality",
            nationality_totals,
            total,
            source_file,
        )
    )
    metrics.extend(
        calculate_top_aggregate_share_metrics(
            date_value,
            "residence_status",
            "foreign_residents_by_status",
            status_totals,
            total,
            source_file,
        )
    )
    metrics.append(
        ForeignResidentMetric(
            date=date_value,
            metric_key="long_term_settlement_share",
            dimension="residence_status_group",
            dimension_value="long_term_settlement",
            value=float(long_term_total / total * 100),
            unit="%",
            source_name=source_name,
            source_file=source_file,
        )
    )
    return metrics


def calculate_top_aggregate_share_metrics(
    date_value: str,
    dimension: str,
    metric_key: str,
    totals: dict[str, float],
    total: float,
    source_file: str,
) -> list[ForeignResidentMetric]:
    metrics: list[ForeignResidentMetric] = []
    for dimension_value, value in sorted(totals.items(), key=lambda item: item[1], reverse=True)[:10]:
        metrics.append(
            ForeignResidentMetric(
                date=date_value,
                metric_key=metric_key,
                dimension=dimension,
                dimension_value=dimension_value,
                value=float(value),
                unit="persons",
                source_name="e-Stat Foreign Residents",
                source_file=source_file,
            )
        )
        metrics.append(
            ForeignResidentMetric(
                date=date_value,
                metric_key=f"{metric_key}_share",
                dimension=dimension,
                dimension_value=dimension_value,
                value=float(value / total * 100),
                unit="%",
                source_name="e-Stat Foreign Residents",
                source_file=source_file,
            )
        )
    return metrics


def calculate_foreign_resident_change_metrics(metrics: list[ForeignResidentMetric]) -> list[ForeignResidentMetric]:
    totals = [
        item
        for item in metrics
        if item.metric_key == "foreign_residents_total"
        and item.dimension == "total"
        and item.dimension_value == "all"
        and item.value is not None
    ]
    if len(totals) < 2:
        return []

    source_name = "e-Stat Foreign Residents"
    value_by_date = {item.date: float(item.value or 0) for item in totals}
    source_by_date = {item.date: item.source_file for item in totals}
    out: list[ForeignResidentMetric] = []

    for current_date, current_value in sorted(value_by_date.items()):
        current = pd.Timestamp(current_date)
        previous_year = (current - pd.DateOffset(years=1)).date().isoformat()
        previous_five_years = (current - pd.DateOffset(years=5)).date().isoformat()
        if previous_year in value_by_date and value_by_date[previous_year]:
            yoy = (current_value / value_by_date[previous_year] - 1) * 100
            out.append(
                ForeignResidentMetric(
                    current_date,
                    "foreign_residents_yoy",
                    "total",
                    "all",
                    float(yoy),
                    "%",
                    source_name,
                    source_by_date.get(current_date, ""),
                )
            )
        if previous_five_years in value_by_date and value_by_date[previous_five_years]:
            cagr = (current_value / value_by_date[previous_five_years]) ** (1 / 5) - 1
            out.append(
                ForeignResidentMetric(
                    current_date,
                    "foreign_residents_cagr_5y",
                    "total",
                    "all",
                    float(cagr * 100),
                    "%",
                    source_name,
                    source_by_date.get(current_date, ""),
                )
            )
    return out


def clean_foreign_resident_file(path: Path) -> list[ForeignResidentObservation]:
    xls = pd.ExcelFile(path)
    detail_sheet = find_detail_sheet(xls.sheet_names)
    if detail_sheet is None:
        return []

    year, month = parse_period_from_sheet(detail_sheet)
    period_date = month_end(year, month)
    columns = {"国籍・地域", "在留資格", "性別", "年齢（５歳階級）", "年齢", "都道府県", "在留外国人数"}
    df = pd.read_excel(path, sheet_name=detail_sheet, dtype=str, usecols=lambda name: name in columns)
    logger.info("读取在留外国人明细 sheet=%s：%d 行", detail_sheet, len(df))

    required = {"国籍・地域", "在留資格", "都道府県", "在留外国人数"}
    if not required.issubset(set(df.columns)):
        return []

    out: list[ForeignResidentObservation] = []
    for row in df.itertuples(index=False):
        item = dict(zip(df.columns, row, strict=False))
        value = parse_int(item.get("在留外国人数"))
        if value is None:
            continue

        nationality_code, nationality = split_code_label(item.get("国籍・地域"))
        status_code, residence_status = split_code_label(item.get("在留資格"))
        prefecture_code, prefecture = split_code_label(item.get("都道府県"))
        _, gender = split_code_label(item.get("性別"))
        _, age_group = split_code_label(item.get("年齢（５歳階級）"))
        age = clean_text(item.get("年齢"))

        out.append(
            ForeignResidentObservation(
                date=period_date,
                year=year,
                month=month,
                nationality=nationality,
                nationality_code=nationality_code,
                residence_status=residence_status,
                residence_status_code=status_code,
                prefecture=prefecture,
                prefecture_code=prefecture_code,
                gender=gender,
                age_group=age_group,
                age=age,
                value=value,
                unit="persons",
                source_name="e-Stat Foreign Residents",
                source_file=str(path),
            )
        )
    return out


def find_detail_sheet(sheet_names: list[str]) -> str | None:
    for name in sheet_names:
        if re.search(r"令和\d+年\d+月末", name):
            return name
    return None


def parse_period_from_sheet(sheet_name: str) -> tuple[int, int]:
    match = re.search(r"令和(\d+)年(\d+)月末", sheet_name)
    if not match:
        raise ValueError(f"Cannot parse Japanese era period from sheet name: {sheet_name}")
    year = 2018 + int(match.group(1))
    month = int(match.group(2))
    return year, month


def split_code_label(value: object) -> tuple[str | None, str]:
    text = clean_text(value) or ""
    if not text:
        return None, ""
    for separator in ("：", ":"):
        if separator in text:
            code, label = text.split(separator, 1)
            return code.strip() or None, label.strip()
    return None, text.strip()


def clean_text(value: object) -> str | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def parse_int(value: object) -> int | None:
    text = clean_text(value)
    if text is None:
        return None
    try:
        return int(float(text.replace(",", "")))
    except ValueError:
        return None


def month_end(year: int, month: int) -> str:
    return date(year, month, calendar.monthrange(year, month)[1]).isoformat()


def calculate_foreign_resident_metrics(
    observations: list[ForeignResidentObservation],
) -> list[ForeignResidentMetric]:
    if not observations:
        return []

    frame = pd.DataFrame([item.__dict__ for item in observations])
    return calculate_foreign_resident_metrics_from_frame(frame)


def calculate_foreign_resident_metrics_from_frame(frame: pd.DataFrame) -> list[ForeignResidentMetric]:
    if frame.empty:
        return []
    metrics: list[ForeignResidentMetric] = []
    source_name = "e-Stat Foreign Residents"

    totals = frame.groupby("date", as_index=False)["value"].sum().sort_values("date")
    source_by_date = frame.groupby("date")["source_file"].first().to_dict()

    for row in totals.itertuples(index=False):
        metrics.append(
            ForeignResidentMetric(
                date=row.date,
                metric_key="foreign_residents_total",
                dimension="total",
                dimension_value="all",
                value=float(row.value),
                unit="persons",
                source_name=source_name,
                source_file=str(source_by_date.get(row.date, "")),
            )
        )

    metrics.extend(calculate_change_metrics(totals, source_by_date, source_name))
    metrics.extend(calculate_share_metrics(frame, "nationality", "foreign_residents_by_nationality", source_by_date))
    metrics.extend(calculate_share_metrics(frame, "residence_status", "foreign_residents_by_status", source_by_date))
    metrics.extend(calculate_long_term_share(frame, source_by_date))
    return metrics


def calculate_change_metrics(
    totals: pd.DataFrame,
    source_by_date: dict[str, str],
    source_name: str,
) -> list[ForeignResidentMetric]:
    metrics: list[ForeignResidentMetric] = []
    value_by_date = dict(zip(totals["date"], totals["value"], strict=False))
    for current_date, current_value in value_by_date.items():
        current = pd.Timestamp(current_date)
        previous_year = (current - pd.DateOffset(years=1)).date().isoformat()
        previous_five_years = (current - pd.DateOffset(years=5)).date().isoformat()
        if previous_year in value_by_date and value_by_date[previous_year]:
            yoy = (current_value / value_by_date[previous_year] - 1) * 100
            metrics.append(
                ForeignResidentMetric(
                    current_date,
                    "foreign_residents_yoy",
                    "total",
                    "all",
                    float(yoy),
                    "%",
                    source_name,
                    str(source_by_date.get(current_date, "")),
                )
            )
        if previous_five_years in value_by_date and value_by_date[previous_five_years]:
            cagr = (current_value / value_by_date[previous_five_years]) ** (1 / 5) - 1
            metrics.append(
                ForeignResidentMetric(
                    current_date,
                    "foreign_residents_cagr_5y",
                    "total",
                    "all",
                    float(cagr * 100),
                    "%",
                    source_name,
                    str(source_by_date.get(current_date, "")),
                )
            )
    return metrics


def calculate_share_metrics(
    frame: pd.DataFrame,
    dimension: str,
    metric_key: str,
    source_by_date: dict[str, str],
) -> list[ForeignResidentMetric]:
    metrics: list[ForeignResidentMetric] = []
    grouped = frame.groupby(["date", dimension], as_index=False)["value"].sum()
    totals = frame.groupby("date")["value"].sum().to_dict()
    shares: dict[tuple[str, str], float] = {}
    for date_value, group in grouped.groupby("date"):
        ranked = group.sort_values("value", ascending=False).head(10)
        total = totals.get(date_value)
        if not total:
            continue
        for row in ranked.itertuples(index=False):
            metrics.append(
                ForeignResidentMetric(
                    date=date_value,
                    metric_key=metric_key,
                    dimension=dimension,
                    dimension_value=str(getattr(row, dimension)),
                    value=float(row.value),
                    unit="persons",
                    source_name="e-Stat Foreign Residents",
                    source_file=str(source_by_date.get(date_value, "")),
                )
            )
            share = float(row.value / total * 100)
            dimension_value = str(getattr(row, dimension))
            shares[(date_value, dimension_value)] = share
            metrics.append(
                ForeignResidentMetric(
                    date=date_value,
                    metric_key=f"{metric_key}_share",
                    dimension=dimension,
                    dimension_value=dimension_value,
                    value=share,
                    unit="%",
                    source_name="e-Stat Foreign Residents",
                    source_file=str(source_by_date.get(date_value, "")),
                )
            )
    for (date_value, dimension_value), current_share in shares.items():
        previous_date = (pd.Timestamp(date_value) - pd.DateOffset(years=5)).date().isoformat()
        previous_share = shares.get((previous_date, dimension_value))
        if previous_share is None:
            continue
        metrics.append(
            ForeignResidentMetric(
                date=date_value,
                metric_key=f"{metric_key}_share_change_5y",
                dimension=dimension,
                dimension_value=dimension_value,
                value=float(current_share - previous_share),
                unit="percentage_points",
                source_name="e-Stat Foreign Residents",
                source_file=str(source_by_date.get(date_value, "")),
            )
        )
    return metrics


def calculate_long_term_share(
    frame: pd.DataFrame,
    source_by_date: dict[str, str],
) -> list[ForeignResidentMetric]:
    metrics: list[ForeignResidentMetric] = []
    temp = frame.copy()
    temp["is_long_term"] = temp["residence_status"].apply(is_long_term_status)
    grouped = temp.groupby(["date", "is_long_term"], as_index=False)["value"].sum()
    totals = temp.groupby("date")["value"].sum().to_dict()
    for row in grouped.itertuples(index=False):
        if not bool(row.is_long_term):
            continue
        total = totals.get(row.date)
        if not total:
            continue
        metrics.append(
            ForeignResidentMetric(
                date=row.date,
                metric_key="long_term_settlement_share",
                dimension="residence_status_group",
                dimension_value="long_term_settlement",
                value=float(row.value / total * 100),
                unit="%",
                source_name="e-Stat Foreign Residents",
                source_file=str(source_by_date.get(row.date, "")),
            )
        )
    return metrics


def is_long_term_status(status: str) -> bool:
    return any(keyword in status for keyword in LONG_TERM_STATUS_KEYWORDS)


def clean_foreign_worker_metrics(source_dir: Path) -> list[ForeignWorkerMetric]:
    files = sorted(source_dir.glob("mhlw_foreign_workers*.xlsx"))
    out: list[ForeignWorkerMetric] = []
    for path in files:
        logger.info("清洗外国人雇用状况文件：%s", path)
        out.extend(clean_mhlw_foreign_worker_file(path))
    return out


def clean_mhlw_foreign_worker_file(path: Path) -> list[ForeignWorkerMetric]:
    source_name = "MHLW Foreign Workers"
    out: list[ForeignWorkerMetric] = []
    xls = pd.ExcelFile(path)
    period_date = parse_mhlw_period(path, xls)

    if "別表４" in xls.sheet_names:
        industry = pd.read_excel(path, sheet_name="別表４", header=None)
        total = parse_float_like(industry.iloc[4, 10])
        rows = []
        for _, row in industry.iloc[5:].iterrows():
            code = clean_text(row.iloc[1]) if len(row) > 1 else None
            label = clean_text(row.iloc[2]) if len(row) > 2 else None
            value = parse_float_like(row.iloc[10]) if len(row) > 10 else None
            if code is None or label is None or value is None:
                continue
            rows.append((label, value))
        rows = sorted(rows, key=lambda item: item[1], reverse=True)[:10]
        for label, value in rows:
            out.append(ForeignWorkerMetric(period_date, "foreign_workers_by_industry", "industry", label, value, "persons", source_name, str(path)))
            if total:
                out.append(ForeignWorkerMetric(period_date, "foreign_workers_by_industry_share", "industry", label, value / total * 100, "%", source_name, str(path)))
        if total:
            top3 = sum(value for _, value in rows[:3]) / total * 100
            out.append(ForeignWorkerMetric(period_date, "industry_concentration_top3", "total", "top3_industries", top3, "%", source_name, str(path)))

    if "別表２" in xls.sheet_names:
        pref = pd.read_excel(path, sheet_name="別表２", header=None)
        total = parse_float_like(pref.iloc[5, 6])
        rows = []
        for _, row in pref.iloc[6:].iterrows():
            label = clean_text(row.iloc[1]) if len(row) > 1 else None
            value = parse_float_like(row.iloc[6]) if len(row) > 6 else None
            if label is None or value is None:
                continue
            rows.append((label, value))
        rows = sorted(rows, key=lambda item: item[1], reverse=True)[:10]
        for label, value in rows:
            out.append(ForeignWorkerMetric(period_date, "foreign_workers_by_prefecture", "prefecture", label, value, "persons", source_name, str(path)))
            if total:
                out.append(ForeignWorkerMetric(period_date, "foreign_workers_by_prefecture_share", "prefecture", label, value / total * 100, "%", source_name, str(path)))
    return out


def clean_foreign_wage_metrics(
    source_dir: Path,
    macro_observations: list[Observation],
) -> list[ForeignWageMetric]:
    files = sorted(source_dir.glob("estat_foreign_wages*.xlsx"))
    out: list[ForeignWageMetric] = []
    for path in files:
        logger.info("清洗外国人工资文件：%s", path)
        out.extend(clean_foreign_wage_file(path, macro_observations))
    return out


def clean_foreign_wage_file(path: Path, macro_observations: list[Observation]) -> list[ForeignWageMetric]:
    source_name = "e-Stat Foreign Wages"
    out: list[ForeignWageMetric] = []
    df = pd.read_excel(path, sheet_name=0, header=None)
    period_date = parse_foreign_wage_period(df)
    cpi_yoy = latest_macro_value(macro_observations, "cpi_yoy")

    current_group: str | None = None
    group_rows: list[tuple[str, float]] = []
    for _, row in df.iloc[9:].iterrows():
        raw_label = clean_text(row.iloc[2]) if len(row) > 2 else None
        if not raw_label:
            continue
        label = normalize_wage_label(raw_label)
        wage = parse_float_like(row.iloc[7]) if len(row) > 7 else None
        worker_count = parse_float_like(row.iloc[10]) if len(row) > 10 else None
        if "産業・企業規模計" in label:
            current_group = label.replace("産業・企業規模計", "").strip() or "外国人労働者"
            if wage is not None:
                group_rows.append((current_group, wage))
                out.append(ForeignWageMetric(period_date, "foreign_nominal_wage", "worker_group", current_group, wage, "thousand_jpy", source_name, str(path)))
                if cpi_yoy is not None:
                    out.append(ForeignWageMetric(period_date, "foreign_wage_minus_cpi", "worker_group", current_group, wage - cpi_yoy, "approx_index_gap", source_name, str(path)))
            if worker_count is not None:
                out.append(ForeignWageMetric(period_date, "foreign_wage_survey_workers", "worker_group", current_group, worker_count * 10, "persons", source_name, str(path)))
            continue
        if current_group and is_major_industry_label(label) and wage is not None:
            out.append(ForeignWageMetric(period_date, "foreign_nominal_wage_by_industry", "industry", f"{current_group} / {label}", wage, "thousand_jpy", source_name, str(path)))

    if group_rows:
        baseline = dict(group_rows).get("外国人労働者")
        if baseline:
            for group, wage in group_rows:
                out.append(ForeignWageMetric(period_date, "foreign_wage_gap_vs_all_foreign_workers", "worker_group", group, wage - baseline, "thousand_jpy", source_name, str(path)))
    return out


def normalize_wage_label(value: str) -> str:
    return re.sub(r"\s+", "", value.replace("\u3000", " ").replace("\n", " ")).strip()


def parse_mhlw_period(path: Path, xls: pd.ExcelFile) -> str:
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=4)
        text = " ".join(str(value) for value in df.fillna("").to_numpy().ravel())
        match = re.search(r"令和(\d+)年10月末", text)
        if match:
            year = 2018 + int(match.group(1))
            return date(year, 10, 31).isoformat()
    return date.today().isoformat()


def parse_foreign_wage_period(df: pd.DataFrame) -> str:
    text = " ".join(str(value) for value in df.head(3).fillna("").to_numpy().ravel())
    match = re.search(r"令和(\d+)年", text)
    if match:
        year = 2018 + int(match.group(1))
        return date(year, 12, 31).isoformat()
    return date.today().isoformat()


def is_major_industry_label(label: str) -> bool:
    return bool(re.match(r"^[A-ZＡ-Ｚ]", label)) and "産業・企業規模計" not in label


def parse_float_like(value: object) -> float | None:
    text = clean_text(value)
    if text is None or text in {"X", "-", "***"}:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def latest_macro_value(observations: list[Observation], series_key: str) -> float | None:
    matches = [item for item in observations if item.series_key == series_key and item.value is not None]
    if not matches:
        return None
    return sorted(matches, key=lambda item: item.date)[-1].value


def main() -> None:
    from src.config import load_config

    config = load_config()
    observations, metrics, worker_metrics, wage_metrics = clean_foreign_residents(config.paths.raw_dir)
    print(f"foreign resident observations: {len(observations)}")
    print(f"foreign resident metrics: {len(metrics)}")
    print(f"foreign worker metrics: {len(worker_metrics)}")
    print(f"foreign wage metrics: {len(wage_metrics)}")


if __name__ == "__main__":
    main()
