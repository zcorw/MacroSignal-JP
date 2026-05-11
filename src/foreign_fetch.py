from __future__ import annotations

import logging
import re
import tempfile
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup

from src.config import AppConfig, load_config
from src.fetch_data import DownloadedFile, discover_estat_download_urls, download_file, extract_stat_id

logger = logging.getLogger(__name__)

ESTAT_FOREIGN_RESIDENTS_TABLE_DATA = (
    "https://www.e-stat.go.jp/stat-search/file-download?fileKind=0&statInfId=000040292372"
)
MHLW_FOREIGN_WORKERS_2024 = "https://www.mhlw.go.jp/content/11655000/001389472.xlsx"
ESTAT_FOREIGN_WAGES_2024 = "https://www.e-stat.go.jp/stat-search/file-download?fileKind=4&statInfId=000040247905"
ESTAT_FOREIGN_RESIDENTS_SEARCH = (
    "https://www.e-stat.go.jp/stat-search/files?cycle=1&layout=dataset"
    "&toukei=00250012&tstat=000001018034&tclass1=000001060399"
)
MHLW_FOREIGN_WORKERS_INDEX = (
    "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/koyou_roudou/koyou/"
    "gaikokujin/gaikokujin-koyou/06.html"
)
ESTAT_FOREIGN_WAGES_SEARCH = (
    "https://www.e-stat.go.jp/stat-search/files?cycle=0&layout=dataset"
    "&toukei=00450091&tstat=000001011429"
    "&tclass1=000001224440&tclass2=000001225782&tclass3=000001225791"
)


def download_foreign_sources(config: AppConfig) -> list[DownloadedFile]:
    raw_dir = config.paths.raw_dir / "foreign_residents"
    raw_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": config.download.user_agent})

    return [
        download_latest_foreign_residents(
            session,
            raw_dir / "estat_foreign_residents_latest.xlsx",
            config,
        ),
        download_latest_mhlw_foreign_workers(
            session,
            raw_dir / "mhlw_foreign_workers_latest.xlsx",
            config,
        ),
        download_latest_foreign_wages(
            session,
            raw_dir / "estat_foreign_wages_latest.xlsx",
            config,
        ),
    ]


def download_latest_foreign_residents(session: requests.Session, path: Path, config: AppConfig) -> DownloadedFile:
    source_name = "e-Stat Foreign Residents"
    try:
        candidates = discover_estat_download_urls(session, ESTAT_FOREIGN_RESIDENTS_SEARCH, "0", config)
        logger.info("%s 动态发现候选下载链接 %d 个", source_name, len(candidates))
        for url in candidates[:20]:
            response = session.get(url, timeout=config.download.timeout_seconds)
            response.raise_for_status()
            if not validate_foreign_residents_file(response.content):
                logger.info("%s 跳过非明细候选：%s", source_name, url)
                continue
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(response.content)
            stat_id = extract_stat_id(url)
            return DownloadedFile(source_name, "success", path, stat_id, f"动态发现并下载成功：{stat_id}")
    except Exception:
        logger.exception("%s 动态发现失败，使用备用链接", source_name)
    return download_file(session, ESTAT_FOREIGN_RESIDENTS_TABLE_DATA, path, source_name, config)


def download_latest_mhlw_foreign_workers(session: requests.Session, path: Path, config: AppConfig) -> DownloadedFile:
    source_name = "MHLW Foreign Workers"
    try:
        page_url = discover_latest_mhlw_result_page(session, config)
        xlsx_url = discover_mhlw_xlsx_url(session, page_url, config)
        result = download_file(session, xlsx_url, path, source_name, config)
        if result.status == "success":
            return DownloadedFile(source_name, result.status, result.path, page_url, f"动态发现并下载成功：{xlsx_url}")
    except Exception:
        logger.exception("%s 动态发现失败，使用备用链接", source_name)
    return download_file(session, MHLW_FOREIGN_WORKERS_2024, path, source_name, config)


def download_latest_foreign_wages(session: requests.Session, path: Path, config: AppConfig) -> DownloadedFile:
    source_name = "e-Stat Foreign Wages"
    try:
        candidates = discover_estat_download_urls(session, ESTAT_FOREIGN_WAGES_SEARCH, "4", config)
        logger.info("%s 动态发现候选下载链接 %d 个", source_name, len(candidates))
        for url in candidates[:10]:
            response = session.get(url, timeout=config.download.timeout_seconds)
            response.raise_for_status()
            if not validate_foreign_wage_file(response.content):
                logger.info("%s 跳过非工资第1表候选：%s", source_name, url)
                continue
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(response.content)
            stat_id = extract_stat_id(url)
            return DownloadedFile(source_name, "success", path, stat_id, f"动态发现并下载成功：{stat_id}")
    except Exception:
        logger.exception("%s 动态发现失败，使用备用链接", source_name)
    return download_file(session, ESTAT_FOREIGN_WAGES_2024, path, source_name, config)


def discover_latest_mhlw_result_page(session: requests.Session, config: AppConfig) -> str:
    response = session.get(MHLW_FOREIGN_WORKERS_INDEX, timeout=config.download.timeout_seconds)
    response.raise_for_status()
    soup = BeautifulSoup(response.content, "html.parser")
    candidates: list[tuple[int, str]] = []
    for anchor in soup.find_all("a", href=True):
        text = anchor.get_text(" ", strip=True)
        match = re.search(r"令和(\d+)年10月末", text)
        if not match:
            continue
        candidates.append((int(match.group(1)), urljoin(MHLW_FOREIGN_WORKERS_INDEX, anchor["href"])))
    if not candidates:
        raise RuntimeError("MHLW 外国人雇用状况索引页未发现年度结果页")
    return max(candidates, key=lambda item: item[0])[1]


def discover_mhlw_xlsx_url(session: requests.Session, page_url: str, config: AppConfig) -> str:
    response = session.get(page_url, timeout=config.download.timeout_seconds)
    response.raise_for_status()
    soup = BeautifulSoup(response.content, "html.parser")
    for anchor in soup.find_all("a", href=True):
        text = anchor.get_text(" ", strip=True)
        href = anchor["href"]
        if ("XLSX" in text.upper() or href.lower().endswith(".xlsx")) and "表一覧" in text:
            return urljoin(page_url, href)
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        if href.lower().endswith(".xlsx"):
            return urljoin(page_url, href)
    raise RuntimeError("MHLW 年度结果页未发现 XLSX 附件")


def validate_foreign_residents_file(content: bytes) -> bool:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        handle.write(content)
        temp_path = Path(handle.name)
    try:
        workbook = load_workbook(temp_path, read_only=True, data_only=True)
        return any(re.search(r"令和\d+年\d+月末", sheet) and workbook[sheet].max_column >= 7 for sheet in workbook.sheetnames)
    except Exception:
        return False
    finally:
        temp_path.unlink(missing_ok=True)


def validate_foreign_wage_file(content: bytes) -> bool:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        handle.write(content)
        temp_path = Path(handle.name)
    try:
        df = pd.read_excel(temp_path, sheet_name=0, header=None, nrows=12)
        text = " ".join(str(value) for value in df.fillna("").to_numpy().ravel())
        return "外国人労働者" in text and "きまって" in text and "現金給与額" in text
    except Exception:
        return False
    finally:
        temp_path.unlink(missing_ok=True)


def main() -> None:
    config = load_config()
    for result in download_foreign_sources(config):
        path = str(result.path) if result.path else "-"
        print(f"{result.source_name}: {result.status} {path} {result.message}")


if __name__ == "__main__":
    main()
